from __future__ import annotations

import textwrap
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(r"D:\GameDev\ProjectDG")
WORKBOOK = ROOT / "docs" / "DefenseGame_Balance_Skill_Summary.xlsx"
OUTPUT_DIR = ROOT / "outputs" / "monster_range_skill_sync_20260713"
FONT_PATH = Path(r"C:\Windows\Fonts\malgun.ttf")
BOLD_FONT_PATH = Path(r"C:\Windows\Fonts\malgunbd.ttf")


def color_of(color, fallback):
    if color is None:
        return fallback
    if color.type == "rgb" and color.rgb:
        value = color.rgb[-6:]
        if value != "000000":
            return f"#{value}"
    return fallback


def font_for(cell, size=14):
    path = BOLD_FONT_PATH if cell.font.bold and BOLD_FONT_PATH.exists() else FONT_PATH
    return ImageFont.truetype(str(path), size=size) if path.exists() else ImageFont.load_default()


def wrap_lines(draw, value, font, width):
    text = "" if value is None else str(value)
    if text.startswith("="):
        text = text[:60] + ("…" if len(text) > 60 else "")
    lines = []
    for paragraph in text.splitlines() or [""]:
        current = ""
        for char in paragraph:
            trial = current + char
            if current and draw.textbbox((0, 0), trial, font=font)[2] > width:
                lines.append(current)
                current = char
            else:
                current = trial
        lines.append(current)
    return lines or [""]


def merged_bounds(ws, row, col):
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return merged.min_row, merged.max_row, merged.min_col, merged.max_col
    return row, row, col, col


def render_sheet(ws, max_row, max_col, output):
    col_widths = []
    for col in range(1, max_col + 1):
        width = ws.column_dimensions[get_column_letter(col)].width or 10
        col_widths.append(max(72, min(250, int(width * 7.2))))
    row_heights = []
    for row in range(1, max_row + 1):
        height = ws.row_dimensions[row].height or 18
        row_heights.append(max(28, int(height * 1.35)))

    margin = 16
    x_positions = [margin]
    for width in col_widths:
        x_positions.append(x_positions[-1] + width)
    y_positions = [margin]
    for height in row_heights:
        y_positions.append(y_positions[-1] + height)

    image = Image.new("RGB", (x_positions[-1] + margin, y_positions[-1] + margin), "#F3F6FA")
    draw = ImageDraw.Draw(image)
    rendered_merges = set()

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            bounds = merged_bounds(ws, row, col)
            if bounds in rendered_merges:
                continue
            rendered_merges.add(bounds)
            min_row, max_merged_row, min_col, max_merged_col = bounds
            anchor = ws.cell(min_row, min_col)
            x1 = x_positions[min_col - 1]
            x2 = x_positions[max_merged_col]
            y1 = y_positions[min_row - 1]
            y2 = y_positions[max_merged_row]
            fallback = "#FFFFFF"
            if min_row in (1, 15):
                fallback = "#18355E"
            elif min_row == 4:
                fallback = "#2F75B5"
            fill = color_of(anchor.fill.fgColor, fallback)
            draw.rectangle((x1, y1, x2, y2), fill=fill, outline="#9AA9BC", width=1)

            if anchor.value is None:
                continue
            font = font_for(anchor, 16 if min_row in (1, 15) else 13)
            text_color = color_of(anchor.font.color, "#FFFFFF" if min_row in (1, 4, 15) else "#182538")
            lines = wrap_lines(draw, anchor.value, font, max(20, x2 - x1 - 12))
            line_height = font.size + 5 if hasattr(font, "size") else 18
            max_lines = max(1, int((y2 - y1 - 6) / line_height))
            lines = lines[:max_lines]
            block_height = len(lines) * line_height
            y = y1 + max(4, (y2 - y1 - block_height) // 2)
            for line in lines:
                bbox = draw.textbbox((0, 0), line, font=font)
                text_width = bbox[2] - bbox[0]
                x = x1 + 6
                if anchor.alignment.horizontal == "center" or min_row in (1, 4, 15):
                    x = x1 + max(6, (x2 - x1 - text_width) // 2)
                draw.text((x, y), line, font=font, fill=text_color)
                y += line_height

    image.save(output)


def main():
    wb = load_workbook(WORKBOOK, data_only=False)
    render_sheet(wb["몬스터전투스킬"], 21, 16, OUTPUT_DIR / "monster_summary_preview.png")
    render_sheet(wb["몬스터_전투스킬"], 16, 24, OUTPUT_DIR / "monster_detail_preview.png")
    wb.close()
    print(OUTPUT_DIR / "monster_summary_preview.png")
    print(OUTPUT_DIR / "monster_detail_preview.png")


if __name__ == "__main__":
    main()
