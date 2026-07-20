from __future__ import annotations

import hashlib
import json
import os
import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


ROOT = Path(r"D:\GameDev\ProjectDG")
SOURCE = ROOT / "docs" / "DefenseGame_Balance_Skill_Summary.xlsx"
OUTPUT_DIR = ROOT / "outputs" / "monster_range_skill_sync_20260713"
OUTPUT = OUTPUT_DIR / "DefenseGame_Balance_Skill_Summary.xlsx"


MONSTERS = {
    "mob_01": {"range": 2.6, "hp": 940, "attack": 20, "attack_speed": 0.88, "move": 1.15},
    "mob_02": {"range": 1.8, "hp": 42, "attack": 6.56, "attack_speed": 0.82, "move": 1.8},
    "mob_03": {"range": 1.9, "hp": 43.2, "attack": 6.56, "attack_speed": 0.82, "move": 0.9},
    "mob_04": {"range": 2.4, "hp": 1265.6, "attack": 24.514, "attack_speed": 0.8235, "move": 1.1},
    "mob_05": {"range": 3, "hp": 44.4, "attack": 6.56, "attack_speed": 0.82, "move": 1.35},
    "mob_06": {"range": 1.8, "hp": 45.6, "attack": 6.56, "attack_speed": 0.82, "move": 1.2},
    "mob_07": {"range": 3, "hp": 1452, "attack": 27.6, "attack_speed": 0.95, "move": 1.9},
    "mob_08": {"range": 1.5, "hp": 46.8, "attack": 6.56, "attack_speed": 0.902, "move": 0.95},
    "mob_09": {"range": 1.8, "hp": 69.6, "attack": 7.872, "attack_speed": 0.82, "move": 1.12},
}


SKILLS = {
    ("mob_01", 1): {
        "name": "대지 균열", "type": "MassStun",
        "description": "무작위 유닛 2기에 공격력 85% 피해를 주고 짧게 기절시킵니다.",
        "core": "공격력 85%, 무작위 2기, 기절 1.45초", "coefficient": 0.85,
        "scope": "전체 전장 / 대상 2 / 기절 1.45초", "resource": "마나 82 / CD 8.5초",
    },
    ("mob_01", 2): {
        "name": "암석 방벽", "type": "BossFortify",
        "description": "체력을 10% 회복하고 잠시 공격 속도를 높입니다.",
        "core": "최대 HP 10% 회복, 4.5초 강화", "coefficient": 0.1,
        "scope": "자신 / 지속 4.5초 / 범위 -", "resource": "마나 95 / CD 9.5초",
    },
    ("mob_02", 1): {
        "name": "Savage Strike", "type": "DirectDamage",
        "description": "가장 가까운 유닛에게 강한 피해를 줍니다.",
        "core": "공격력 170%, 시전 거리 3m", "coefficient": 1.7,
        "scope": "대상 1 / 지속 - / 시전 3m", "resource": "마나 100 / CD 5초",
    },
    ("mob_03", 1): {
        "name": "Rush", "type": "MoveSpeedBoost",
        "description": "일시적으로 이동속도를 높입니다.",
        "core": "이동속도 +50%, 4초", "coefficient": 0.5,
        "scope": "자신 / 지속 4초 / 범위 -", "resource": "마나 100 / CD 9초",
    },
    ("mob_04", 1): {
        "name": "죽음의 서약", "type": "DeathPact",
        "description": "무작위 유닛 하나를 처형합니다.",
        "core": "무작위 1기 처형, 명시적 전역기", "coefficient": 0,
        "scope": "전체 전장 / 대상 1 / 지속 -", "resource": "마나 100 / CD 14초",
    },
    ("mob_04", 2): {
        "name": "여왕의 속박", "type": "Stun",
        "description": "가장 가까운 유닛을 기절시킵니다.",
        "core": "기절 2.4초, 시전 거리 3m", "coefficient": 0,
        "scope": "대상 1 / 기절 2.4초 / 시전 3m", "resource": "마나 75 / CD 7초",
    },
    ("mob_05", 1): {
        "name": "Crushing Grip", "type": "Stun",
        "description": "유닛 하나를 짧게 기절시킵니다.",
        "core": "기절 1.15초, 시전 거리 3m", "coefficient": 0,
        "scope": "대상 1 / 기절 1.15초 / 시전 3m", "resource": "마나 100 / CD 9초",
    },
    ("mob_06", 1): {
        "name": "Savage Strike", "type": "DirectDamage",
        "description": "가장 가까운 유닛에게 강한 피해를 줍니다.",
        "core": "공격력 170%, 시전 거리 3m", "coefficient": 1.7,
        "scope": "대상 1 / 지속 - / 시전 3m", "resource": "마나 100 / CD 5초",
    },
    ("mob_07", 1): {
        "name": "운석 낙하", "type": "AreaDamage",
        "description": "주변 유닛에게 공격력 125%의 광역 피해를 줍니다.",
        "core": "공격력 125%, 반경 4m", "coefficient": 1.25,
        "scope": "반경 4m / 범위 안 유닛이 있을 때만 발동", "resource": "마나 80 / CD 9.5초",
    },
    ("mob_07", 2): {
        "name": "마나 침식", "type": "ManaBurn",
        "description": "무작위 유닛 3기의 마나를 45% 태웁니다.",
        "core": "마나 45%, 무작위 3기, 명시적 전역기", "coefficient": 0.45,
        "scope": "전체 전장 / 대상 3 / 마나 45% 소각", "resource": "마나 92 / CD 9초",
    },
    ("mob_08", 1): {
        "name": "Rush", "type": "MoveSpeedBoost",
        "description": "일시적으로 이동속도를 높입니다.",
        "core": "이동속도 +50%, 4초", "coefficient": 0.5,
        "scope": "자신 / 지속 4초 / 범위 -", "resource": "마나 100 / CD 9초",
    },
    ("mob_09", 1): {
        "name": "Crushing Grip", "type": "Stun",
        "description": "유닛 하나를 짧게 기절시킵니다.",
        "core": "기절 1.15초, 시전 거리 3m", "coefficient": 0,
        "scope": "대상 1 / 기절 1.15초 / 시전 3m", "resource": "마나 100 / CD 9초",
    },
}


def copy_style(source, target):
    target._style = copy(source._style)
    target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def update_detail_sheet(ws):
    ws["A1"] = "몬스터 전투·스킬 설정 - 런타임 기준"
    ws["A2"] = (
        "2026-07-13 런타임 기준. 원거리 평타 최대 3m, 반격 가능한 유닛이 없으면 추가 접근. "
        "단일 피해·기절 3m, 범위기는 실제 반경, 돌진 3.6m 후 접근, 명시된 보스 전역기만 거리 제한을 무시합니다."
    )
    ws["A2"].alignment = copy(ws["A1"].alignment)

    for row in range(5, ws.max_row + 1):
        monster_id = ws.cell(row, 2).value
        slot = ws.cell(row, 15).value
        if monster_id not in MONSTERS:
            continue

        monster = MONSTERS[monster_id]
        ws.cell(row, 7).value = monster["range"]
        ws.cell(row, 8).value = monster["hp"]
        ws.cell(row, 9).value = monster["attack"]
        ws.cell(row, 10).value = monster["attack_speed"]
        ws.cell(row, 11).value = monster["move"]

        skill = SKILLS.get((monster_id, int(slot))) if isinstance(slot, (int, float)) else None
        if not skill:
            continue
        ws.cell(row, 16).value = skill["name"]
        ws.cell(row, 17).value = skill["type"]
        ws.cell(row, 18).value = skill["description"]
        ws.cell(row, 19).value = skill["core"]
        ws.cell(row, 20).value = skill["coefficient"]
        ws.cell(row, 21).value = f'=IF(OR(Q{row}="DirectDamage",Q{row}="AreaDamage",Q{row}="MassStun"),I{row}*T{row},"")'
        ws.cell(row, 22).value = skill["scope"]
        ws.cell(row, 23).value = skill["resource"]


def update_compact_sheet(ws):
    ws["A2"] = (
        "2026-07-13 런타임·스모크 기준. 원거리 평타는 3m 상한이며 반격 가능한 유닛이 없으면 추가 접근합니다. "
        "9종 모두 AttackHit·SkillHit·FireProjectile·VFX 검증을 통과했습니다."
    )
    for row in range(5, 14):
        monster_id = ws.cell(row, 1).value
        if monster_id not in MONSTERS:
            continue
        ws.cell(row, 5).value = MONSTERS[monster_id]["range"]
        ws.cell(row, 13).value = "통과"
        ws.cell(row, 14).value = "통과"
        ws.cell(row, 15).value = None

    notes = {
        "mob_01": "대지 균열 전역 2기 피해·기절 / 암석 방벽 자가 회복·강화",
        "mob_02": "근접 평타 / 단일 피해 스킬 시전 3m",
        "mob_03": "근접 평타 / 이동속도 강화",
        "mob_04": "죽음의 서약만 명시적 전역 / 여왕의 속박 시전 3m",
        "mob_05": "원거리 평타 3m 상한 + 반격 불가 시 추가 접근 / 기절 시전 3m",
        "mob_06": "근접 평타 / 단일 피해 스킬 시전 3m",
        "mob_07": "원거리 평타 3m 상한 + 반격 불가 시 추가 접근 / 운석 반경 4m·마나 침식 전역",
        "mob_08": "근접 평타 / 이동속도 강화",
        "mob_09": "근접 평타 / 기절 시전 3m",
    }
    for row in range(5, 14):
        monster_id = ws.cell(row, 1).value
        if monster_id in notes:
            ws.cell(row, 16).value = notes[monster_id]

    start_row = 15
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= start_row:
            ws.unmerge_cells(str(merged))
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=16)
    title = ws.cell(start_row, 1)
    copy_style(ws.cell(4, 1), title)
    title.value = "2026-07-13 원거리·스킬 거리 정책"
    title.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(2, 17):
        copy_style(ws.cell(4, col), ws.cell(start_row, col))
    ws.row_dimensions[start_row].height = 24

    policies = [
        ("원거리 평타", "최대 3m. 살아 있는 유닛 중 누구도 반격할 수 없으면 0.12m 여유를 두고 추가 접근"),
        ("단일 피해·기절", "기본 시전 거리 3m, useCustomCastRange 활성"),
        ("범위 피해", "실제 반경 안에 살아 있는 유닛이 있을 때만 사용하며 적중 때 다시 거리 검증"),
        ("돌진", "3.6m 안에서 사용, 최대 1.25m를 0.18초 동안 접근한 뒤 적중"),
        ("전역기", "DeathPact·MassStun·GoldDrain·ManaBurn만 명시적으로 전역 허용"),
        ("몬스터 강화", "기본 5.5m 오라 안의 살아 있는 몬스터만 강화"),
    ]
    for offset, (label, detail) in enumerate(policies, start=1):
        row = start_row + offset
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=16)
        for col in range(1, 17):
            copy_style(ws.cell(5, col), ws.cell(row, col))
            ws.cell(row, col).alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(row, 1).value = label
        ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row, 5).value = detail
        ws.row_dimensions[row].height = 30


def append_history(ws):
    new_row = ws.max_row + 1
    for col in range(1, ws.max_column + 1):
        copy_style(ws.cell(ws.max_row, col), ws.cell(new_row, col))
    ws.cell(new_row, 1).value = "2026-07-13 KST"
    ws.cell(new_row, 2).value = "몬스터 원거리·스킬 거리 정책"
    ws.cell(new_row, 3).value = (
        "원거리 평타 3m 상한 및 반격 불가 시 접근, 단일기 3m·범위기 실반경·돌진 접근·전역기 명시 정책 적용. "
        "Unity 컴파일 0 오류, 9종 보스/애니메이션/VFX 스모크 PASS."
    )
    ws.row_dimensions[new_row].height = max(ws.row_dimensions[ws.max_row - 1].height or 15, 30)


def find_formula_errors(path: Path):
    error_tokens = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
    errors = []
    for data_only in (False, True):
        wb = load_workbook(path, data_only=data_only, read_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and any(token in value for token in error_tokens):
                        errors.append({"sheet": ws.title, "cell": cell.coordinate, "value": value, "data_only": data_only})
        wb.close()
    return errors


def validate(path: Path, expected_sheet_count: int):
    wb = load_workbook(path, data_only=False)
    assert len(wb.sheetnames) == expected_sheet_count
    detail = wb["몬스터_전투스킬"]
    compact = wb["몬스터전투스킬"]

    detail_rows = {(detail.cell(r, 2).value, detail.cell(r, 15).value): r for r in range(5, 17)}
    assert detail.cell(detail_rows[("mob_05", 1)], 7).value == 3
    assert detail.cell(detail_rows[("mob_07", 1)], 7).value == 3
    assert detail.cell(detail_rows[("mob_07", 1)], 16).value == "운석 낙하"
    assert detail.cell(detail_rows[("mob_01", 2)], 16).value == "암석 방벽"
    assert "MassStun" in detail.cell(detail_rows[("mob_01", 1)], 21).value

    compact_rows = {compact.cell(r, 1).value: r for r in range(5, 14)}
    assert compact.cell(compact_rows["mob_05"], 5).value == 3
    assert compact.cell(compact_rows["mob_07"], 5).value == 3
    assert all(compact.cell(r, 13).value == "통과" and compact.cell(r, 14).value == "통과" for r in range(5, 14))
    assert compact.cell(15, 1).value == "2026-07-13 원거리·스킬 거리 정책"
    assert compact.cell(21, 5).value == "기본 5.5m 오라 안의 살아 있는 몬스터만 강화"
    wb.close()

    formula_errors = find_formula_errors(path)
    assert not formula_errors, formula_errors
    return {
        "sheet_count": expected_sheet_count,
        "monster_detail_rows": 12,
        "monster_summary_rows": 9,
        "policy_rows": 6,
        "formula_errors": 0,
    }


def sha256(path: Path):
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(SOURCE, data_only=False)
    expected_sheet_count = len(wb.sheetnames)
    update_detail_sheet(wb["몬스터_전투스킬"])
    update_compact_sheet(wb["몬스터전투스킬"])
    append_history(wb["변경이력"])
    wb.save(OUTPUT)
    wb.close()

    validation = validate(OUTPUT, expected_sheet_count)
    temporary = SOURCE.with_suffix(".xlsx.codex-tmp")
    shutil.copy2(OUTPUT, temporary)
    os.replace(temporary, SOURCE)
    source_validation = validate(SOURCE, expected_sheet_count)

    result = {
        "output": str(OUTPUT),
        "source": str(SOURCE),
        "output_sha256": sha256(OUTPUT),
        "source_sha256": sha256(SOURCE),
        "validation": validation,
        "source_validation": source_validation,
    }
    assert result["output_sha256"] == result["source_sha256"]
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
