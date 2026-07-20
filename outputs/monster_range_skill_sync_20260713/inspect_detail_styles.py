from openpyxl import load_workbook

path = r"D:\GameDev\ProjectDG\docs\DefenseGame_Balance_Skill_Summary.xlsx"
wb = load_workbook(path, data_only=False)
ws = wb["몬스터_전투스킬"]
for row in range(13, 17):
    print("ROW", row, "height", ws.row_dimensions[row].height, "hidden", ws.row_dimensions[row].hidden)
    print("values", [ws.cell(row, col).value for col in (2, 15, 16, 17, 22)])
    print("fills", [(ws.cell(row, col).fill.fgColor.type, ws.cell(row, col).fill.fgColor.rgb, ws.cell(row, col).fill.fgColor.indexed, ws.cell(row, col).fill.fgColor.theme) for col in (1, 2, 16, 24)])
print("merged", [str(x) for x in ws.merged_cells.ranges])
