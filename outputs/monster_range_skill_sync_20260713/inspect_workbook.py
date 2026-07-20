from pathlib import Path
import json

from openpyxl import load_workbook


path = Path(r"D:\GameDev\ProjectDG\docs\DefenseGame_Balance_Skill_Summary.xlsx")
workbook = load_workbook(path, data_only=False)
payload = {
    "sheets": [sheet.title for sheet in workbook.worksheets],
    "monster_sheet": {
        "title": workbook.worksheets[14].title,
        "rows": [
            [workbook.worksheets[14].cell(row, column).value for column in range(1, workbook.worksheets[14].max_column + 1)]
            for row in range(1, workbook.worksheets[14].max_row + 1)
        ],
    },
}
print(json.dumps(payload, ensure_ascii=True, indent=2, default=str))
