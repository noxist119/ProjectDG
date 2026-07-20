from openpyxl import load_workbook

path = r"D:\GameDev\ProjectDG\docs\DefenseGame_Balance_Skill_Summary.xlsx"
wb = load_workbook(path, data_only=False)
ws = wb["변경이력"]
print(ws.max_row, ws.max_column)
for row in ws.iter_rows(min_row=max(1, ws.max_row - 7), max_row=ws.max_row, values_only=True):
    print(row)
