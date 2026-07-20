from pathlib import Path
import json

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"D:\GameDev\ProjectDG\docs\DefenseGame_Balance_Skill_Summary.xlsx")


def main() -> None:
    workbook = load_workbook(WORKBOOK_PATH, data_only=False)
    summary = []
    for sheet in workbook.worksheets:
        summary.append({
            "name": sheet.title,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
        })
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
