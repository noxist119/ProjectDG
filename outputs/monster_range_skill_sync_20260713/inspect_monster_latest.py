from pathlib import Path
import json

from openpyxl import load_workbook


path = Path(r"D:\GameDev\ProjectDG\docs\DefenseGame_Balance_Skill_Summary.xlsx")
workbook = load_workbook(path, data_only=False)
sheet = workbook.worksheets[20]
payload = {
    "title": sheet.title,
    "merged": [str(item) for item in sheet.merged_cells.ranges],
    "freeze": str(sheet.freeze_panes),
    "rows": [
        [sheet.cell(row, column).value for column in range(1, sheet.max_column + 1)]
        for row in range(1, sheet.max_row + 1)
    ],
}
print(json.dumps(payload, ensure_ascii=True, indent=2, default=str))
