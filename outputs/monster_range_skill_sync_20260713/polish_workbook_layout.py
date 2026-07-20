from __future__ import annotations

import hashlib
import os
import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"D:\GameDev\ProjectDG")
SOURCE = ROOT / "docs" / "DefenseGame_Balance_Skill_Summary.xlsx"
OUTPUT = ROOT / "outputs" / "monster_range_skill_sync_20260713" / "DefenseGame_Balance_Skill_Summary.xlsx"


def digest(path):
    value = hashlib.sha256()
    with open(path, "rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            value.update(chunk)
    return value.hexdigest()


def validate(path):
    wb = load_workbook(path, data_only=False)
    ws = wb["몬스터전투스킬"]
    assert ws["E9"].value == 3
    assert ws["E11"].value == 3
    assert all(ws.cell(row, 13).value == "통과" and ws.cell(row, 14).value == "통과" for row in range(5, 14))
    assert all(ws.row_dimensions[row].height >= 44 for row in range(5, 14))
    assert all(ws.cell(row, 16).alignment.wrap_text for row in range(5, 14))
    assert ws["A15"].value == "2026-07-13 원거리·스킬 거리 정책"
    errors = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and any(token in cell.value for token in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")):
                    errors.append((sheet.title, cell.coordinate, cell.value))
    wb.close()
    assert not errors, errors


wb = load_workbook(SOURCE, data_only=False)
ws = wb["몬스터전투스킬"]
ws.row_dimensions[2].height = 34
ws["A2"].alignment = copy(ws["A2"].alignment)
ws["A2"].alignment = ws["A2"].alignment.copy(wrap_text=True, vertical="center")
for row in range(5, 14):
    ws.row_dimensions[row].height = 44
    for col in range(1, 17):
        cell = ws.cell(row, col)
        cell.alignment = cell.alignment.copy(vertical="center", wrap_text=(col == 16 or cell.alignment.wrap_text))
wb.save(OUTPUT)
wb.close()

validate(OUTPUT)
temporary = SOURCE.with_suffix(".xlsx.codex-tmp")
shutil.copy2(OUTPUT, temporary)
os.replace(temporary, SOURCE)
validate(SOURCE)
assert digest(OUTPUT) == digest(SOURCE)
print(digest(SOURCE))
