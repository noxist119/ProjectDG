from __future__ import annotations

import json
import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from update_balance_9point import ROOT, SOURCE, OUTPUT_DIR, OUTPUT, LATEST_SHEET, VALIDATION_SHEET


LATEST_ROWS = [
    ["운명카드", "3장 선택 구성", "생존 1 + 전투 1 + 성장/위험 1 보장 후 위치 셔플", "DefenseGameController.cs", "적용", "18장 완전 랜덤의 동일 역할 몰림 제거"],
    ["운명카드", "다음 웨이브 반동", "적 수 +65% -> +50%", "DefenseGameController.cs", "적용", "악마와의 거래는 유지하고 R9~R10 급사 편차 완화"],
    ["운명카드", "경제 페널티", "금단 50%->40% / 에픽 35%->30% / 방벽 30%->25%", "DefenseGameController.cs", "적용", "사용 후 경제 회복 불가능 구간 완화"],
    ["운명카드", "도박사의 판 실패", "HP -3 -> -2", "DefenseGameController.cs", "적용", "70% 성공 카드 실패가 즉시 패배로 직결되는 빈도 감소"],
    ["소형상점", "3개 상품 역할 분리", "합성 부스터 + 소환 쿠폰(위기 시 의무병) + 무작위 1개", "RunShopSystem.cs", "적용", "합성/경제/상황대응을 매번 비교 가능"],
    ["소형상점", "소환 쿠폰", "이번 판 소환비 5% -> 8% 할인", "RunShopSystem.cs", "적용", "즉시 소환과 장기 할인 사이 기회비용 강화"],
    ["소형상점", "현장 의무병", "유닛 HP 25% + 방어선 HP 절반 이하일 때 HP 1 회복", "RunShopSystem.cs", "적용", "위기 때만 생존 가치가 생기는 조건부 회복"],
    ["검증", "5판 안전 실측", "96배 -> 40배 / fixed 0.025 / max 0.33 / 고정 시드", "DefenseGameBatchPlaytest.cs", "적용", "AttackHit·SkillHit·투사체 이벤트 누락 방지"],
]


def clone_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format


def clone_row_style(ws, source_row: int, target_row: int, columns: int = 6) -> None:
    for col in range(1, columns + 1):
        clone_cell_style(ws.cell(source_row, col), ws.cell(target_row, col))
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def write_latest_section(ws) -> None:
    marker = "9점 보완 조정 (2026-07-11)"
    marker_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == marker:
            marker_row = row
            break

    if marker_row is None:
        marker_row = ws.max_row + 2
        ws.merge_cells(start_row=marker_row, start_column=1, end_row=marker_row, end_column=6)
        clone_cell_style(ws["A1"], ws.cell(marker_row, 1))
        ws.cell(marker_row, 1, marker)
        ws.row_dimensions[marker_row].height = 26

    header_row = marker_row + 1
    clone_row_style(ws, 4, header_row)
    headers = ["분류", "항목", "최신 설정", "코드 근거", "상태", "메모"]
    for col, value in enumerate(headers, 1):
        ws.cell(header_row, col, value)

    for index, values in enumerate(LATEST_ROWS):
        target_row = header_row + 1 + index
        clone_row_style(ws, 5 if index % 2 == 0 else 6, target_row)
        for col, value in enumerate(values, 1):
            ws.cell(target_row, col, value)
            ws.cell(target_row, col).alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[target_row].height = 44

    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 15)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 24)
    ws.column_dimensions["C"].width = max(ws.column_dimensions["C"].width or 0, 48)
    ws.column_dimensions["D"].width = max(ws.column_dimensions["D"].width or 0, 35)
    ws.column_dimensions["E"].width = max(ws.column_dimensions["E"].width or 0, 12)
    ws.column_dimensions["F"].width = max(ws.column_dimensions["F"].width or 0, 50)


def upsert_validation(ws, item: str, values: list) -> None:
    target_row = None
    for row in range(4, ws.max_row + 1):
        if ws.cell(row, 1).value == item:
            target_row = row
            break
    if target_row is None:
        target_row = ws.max_row + 1
        clone_row_style(ws, 4 if target_row % 2 == 0 else 5, target_row)
    for col, value in enumerate(values, 1):
        ws.cell(target_row, col, value)
        ws.cell(target_row, col).alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[target_row].height = 42


def scan_formula_errors(wb) -> list[str]:
    errors = []
    markers = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and any(marker in value for marker in markers):
                    errors.append(f"{ws.title}!{cell.coordinate}:{value}")
    return errors


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SOURCE, OUTPUT)
    wb = load_workbook(OUTPUT)
    latest = wb[LATEST_SHEET]
    validation = wb[VALIDATION_SHEET]

    write_latest_section(latest)
    upsert_validation(
        validation,
        "9점 보완 C# 빌드",
        ["9점 보완 C# 빌드", "PASS / 경고 0 / 오류 0", "Assembly-CSharp.csproj --no-restore", "통과", "운명카드·상점·테스트 하네스", "전체 컴파일"],
    )
    upsert_validation(
        validation,
        "9점 보완 Play Mode 스모크",
        ["9점 보완 Play Mode 스모크", "PASS / runtimeErrors 0", "BatchPlaytestResults/DefenseGame_PlayModeSmoke.json", "통과", "Safe Area·hero_55~57·애니메이션·VFX", "세로 UI/신규 초월"],
    )
    upsert_validation(
        validation,
        "9점 보완 인간형 3전략 5판",
        ["9점 보완 인간형 3전략 5판", "실측 진행 중", "BatchPlaytestResults/DefenseGame_Playtest5_Human3.json", "대기", "목표 R10 65~75% / 보스 잔여 HP 10~25%", "40배 안전속도·고정 시드"],
    )

    errors = scan_formula_errors(wb)
    wb.save(OUTPUT)

    check = load_workbook(OUTPUT, data_only=False)
    payload = {
        "output": str(OUTPUT),
        "sheet_count": len(check.sheetnames),
        "latest_range": [
            [check[LATEST_SHEET].cell(row, col).value for col in range(1, 7)]
            for row in range(check[LATEST_SHEET].max_row - len(LATEST_ROWS) - 1, check[LATEST_SHEET].max_row + 1)
        ],
        "validation_tail": [
            [check[VALIDATION_SHEET].cell(row, col).value for col in range(1, 7)]
            for row in range(max(1, check[VALIDATION_SHEET].max_row - 4), check[VALIDATION_SHEET].max_row + 1)
        ],
        "formula_errors": errors,
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
