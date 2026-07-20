from __future__ import annotations

import json
import shutil
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[2]
SOURCE = ROOT / "docs" / "DefenseGame_Balance_Skill_Summary.xlsx"
OUTPUT_DIR = Path(__file__).resolve().parent
OUTPUT = OUTPUT_DIR / "DefenseGame_Balance_Skill_Summary.xlsx"
LATEST_SHEET = "Latest_2026-07-11"
VALIDATION_SHEET = "Validation_2026-07-11"


def compact_sheet(ws, max_rows: int = 40, max_cols: int = 10) -> dict:
    rows = []
    for row in ws.iter_rows(
        min_row=1,
        max_row=min(ws.max_row, max_rows),
        min_col=1,
        max_col=min(ws.max_column, max_cols),
    ):
        values = [cell.value for cell in row]
        if any(value not in (None, "") for value in values):
            rows.append(
                {
                    "row": row[0].row,
                    "values": values,
                    "style_ids": [cell.style_id for cell in row],
                }
            )
    return {
        "title": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "freeze_panes": str(ws.freeze_panes or ""),
        "rows": rows,
    }


def inspect_workbook(path: Path) -> None:
    wb = load_workbook(path, data_only=False)
    payload = {
        "path": str(path),
        "sheets": [
            {
                "title": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
            }
            for ws in wb.worksheets
        ],
        "relevant": [
            compact_sheet(wb[name])
            for name in (LATEST_SHEET, VALIDATION_SHEET)
            if name in wb.sheetnames
        ],
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2, default=str))


def copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
    if source_row in ws.row_dimensions:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def ensure_sheet(wb, name: str, headers: list[str]):
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(name)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        ws.append(headers)
        header_fill = PatternFill("solid", fgColor="17365D")
        header_font = Font(color="FFFFFF", bold=True)
        bottom = Side(style="thin", color="9EB6CE")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=bottom)
        ws.row_dimensions[1].height = 24
    return ws


def append_unique(ws, key: str, values: list) -> None:
    for row in ws.iter_rows(min_row=2, max_col=1):
        if row[0].value == key:
            for col, value in enumerate(values, 1):
                ws.cell(row[0].row, col, value)
            return

    target_row = ws.max_row + 1
    if ws.max_row >= 2:
        copy_row_style(ws, ws.max_row, target_row, len(values))
    for col, value in enumerate(values, 1):
        ws.cell(target_row, col, value)


def update_workbook() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SOURCE, OUTPUT)
    wb = load_workbook(OUTPUT)

    latest = ensure_sheet(
        wb,
        LATEST_SHEET,
        ["키", "구분", "대상", "변경 전", "변경 후", "의도", "코드 근거"],
    )
    changes = [
        [
            "9POINT-FATE-DRAFT",
            "운명카드",
            "3장 랜덤 제시",
            "18장 중 완전 랜덤 3장",
            "생존 1 + 전투 1 + 성장/위험 1 보장 후 위치 셔플",
            "동일 역할 카드 몰림을 없애고 위기 순간 선택의 질을 보장",
            "DefenseGameController.EnsureFateCardChoices",
        ],
        [
            "9POINT-FATE-BACKLASH",
            "운명카드",
            "다음 웨이브 적 수 반동",
            0.65,
            0.50,
            "악마와의 거래 감각은 유지하면서 R9~R10 급사 편차 완화",
            "DefenseGameController.fateCardBacklashMonsterCountMultiplier",
        ],
        [
            "9POINT-FATE-TAX",
            "운명카드",
            "금단/에픽/방벽 소환비 페널티",
            "50% / 35% / 30%",
            "40% / 30% / 25%",
            "카드 사용 후 회복 불가능한 경제 경직 완화",
            "DefenseGameController fateCard*CostPenalty",
        ],
        [
            "9POINT-GAMBLER-FAIL",
            "운명카드",
            "도박사의 판 실패 HP",
            -3,
            -2,
            "70% 성공 카드의 실패가 즉시 패배로 직결되는 빈도 감소",
            "DefenseGameController.fateCardGamblerFailLifeCost",
        ],
        [
            "9POINT-MINI-SHOP-MIX",
            "소형상점",
            "R3부터 3개 상품",
            "합성 부스터 + 무작위 2개",
            "합성 부스터 + 소환 쿠폰(위기 시 의무병) + 무작위 1개",
            "합성/경제/상황대응 역할을 분리해 매번 비교 가능한 선택 구성",
            "RunShopSystem.BuildOffers",
        ],
        [
            "9POINT-COUPON",
            "소형상점",
            "소환 쿠폰",
            "소환비 5% 할인",
            "소환비 8% 할인",
            "즉시 유닛 구매와 장기 할인 사이의 기회비용을 체감 가능하게 조정",
            "RunShopSystem.CreateOffer / ApplyOffer",
        ],
        [
            "9POINT-FIELD-MEDIC",
            "소형상점",
            "현장 의무병",
            "유닛 HP 25% 회복",
            "유닛 HP 25% + 방어선 HP 절반 이하일 때 HP 1 회복",
            "위기 때만 생존 가치가 생기는 조건부 회복 상품",
            "RunShopSystem.ApplyOffer",
        ],
        [
            "9POINT-BATCH-SAFETY",
            "검증",
            "5판 배치 테스트 속도",
            "96배 / fixed 0.05 / max 0.75",
            "40배 / fixed 0.025 / max 0.33 / 고정 시드",
            "AttackHit·SkillHit·투사체 물리 이벤트 누락 없는 재현 가능한 실측",
            "DefenseGameBatchPlaytest / DefenseGameBatchTurboOverride",
        ],
    ]
    for row in changes:
        append_unique(latest, row[0], row)

    latest.freeze_panes = "A2"
    latest.sheet_view.showGridLines = False
    widths = {"A": 26, "B": 14, "C": 28, "D": 30, "E": 43, "F": 48, "G": 48}
    for col, width in widths.items():
        latest.column_dimensions[col].width = width
    for row in latest.iter_rows(min_row=2, max_col=7):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in latest[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for key in ("D", "E"):
        for cell in latest[key][1:]:
            if isinstance(cell.value, float):
                cell.number_format = "0%"

    validation = ensure_sheet(
        wb,
        VALIDATION_SHEET,
        ["키", "검증 항목", "결과", "세부", "기준/목표", "근거 파일"],
    )
    validation_rows = [
        [
            "9POINT-CS-BUILD",
            "C# 전체 빌드",
            "PASS",
            "경고 0 / 오류 0",
            "컴파일 오류 0",
            "Assembly-CSharp.csproj --no-restore",
        ],
        [
            "9POINT-PLAYMODE-SMOKE",
            "세로 UI·hero_55~57 프리팹 Play Mode 스모크",
            "PENDING",
            "Unity 재실행 결과 동기화 예정",
            "runtimeErrors 0 / 전체 prefab pass",
            "BatchPlaytestResults/DefenseGame_PlayModeSmoke.json",
        ],
        [
            "9POINT-HUMAN3-5RUN",
            "인간형 3전략 5판 R10 실측",
            "PENDING",
            "40배 안전속도·고정 시드로 재실측 예정",
            "R10 클리어 65~75% / 보스 잔여 HP 10~25%",
            "BatchPlaytestResults/DefenseGame_Playtest5_Human3.json",
        ],
    ]
    for row in validation_rows:
        append_unique(validation, row[0], row)

    validation.freeze_panes = "A2"
    validation.sheet_view.showGridLines = False
    widths = {"A": 28, "B": 42, "C": 14, "D": 48, "E": 44, "F": 55}
    for col, width in widths.items():
        validation.column_dimensions[col].width = width
    for row in validation.iter_rows(min_row=2, max_col=6):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(OUTPUT)
    inspect_workbook(OUTPUT)


if __name__ == "__main__":
    mode = sys.argv[1] if len(sys.argv) > 1 else "inspect"
    if mode == "inspect":
        inspect_workbook(SOURCE)
    elif mode == "update":
        update_workbook()
    else:
        raise SystemExit(f"unknown mode: {mode}")
