from __future__ import annotations

import json
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

import update_balance_9point_v2 as base


RESULT = base.ROOT / "BatchPlaytestResults" / "DefenseGame_Playtest5_Human3.json"


FINAL_LATEST_ROWS = [
    ["경제", "연승 골드 누적", "라운드당 +1 -> +0", "DefenseGameController.cs", "적용", "R10까지 누적되는 무조건 이득을 제거해 상점/소환 기회비용 유지"],
    ["경제", "일반 몬스터 처치 골드", "4+등급x2 -> 3+등급", "MonsterDatabase.cs", "적용", "초반 안정성은 남기고 R7~R10 과잉 소환 재원 축소"],
    ["라운드", "R5~R9 일반 몬스터 완화", "기본 수 -2 -> 기본 수 -1", "RoundManager.cs", "적용", "초반 3소환은 보호하되 보스 전 압박을 절반 복원"],
    ["R10 보스", "첫 보스 스탯", "HP x2.10->x2.75 / 공격 x1.48->x1.78 / 스킬 x1.25->x1.45", "MonsterDatabase.cs", "적용", "일반 라운드는 유지하고 R10 선택 결과만 검증"],
    ["R10 보스", "지원·제어 상한", "지원 4->6기 / 스턴 1->2명 / 광역 x1.45->x1.60", "MonsterDatabase.cs / RoundManager.cs", "적용", "단순 체력벽보다 배치와 상태이상 대응을 요구"],
    ["검증", "R10 실제 종료 판정", "R10 진입 판정 제거 / R10 보스 종료 직후 판정", "DefenseGameBatchPlaytest.cs", "적용", "R10 시작 전 성공 및 R11 진입 오판 제거"],
    ["검증", "전투 중 운명카드 정책", "CanOpen -> 중앙 3장 -> 1장 선택 / 판당 1회", "DefenseGameBatchPlaytest.cs", "적용", "실제 유저의 위기 카드 흐름으로 3전략 검증"],
]


def find_marker_row(ws, marker: str) -> int:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == marker:
            return row
    raise RuntimeError(f"marker not found: {marker}")


def upsert_latest(ws, marker_row: int, values: list) -> None:
    target_row = None
    for row in range(marker_row + 2, ws.max_row + 1):
        if ws.cell(row, 2).value == values[1]:
            target_row = row
            break
    if target_row is None:
        target_row = ws.max_row + 1
        base.clone_row_style(ws, 5 if target_row % 2 else 6, target_row)
    for col, value in enumerate(values, 1):
        ws.cell(target_row, col, value)
        ws.cell(target_row, col).alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[target_row].height = 44


def summarize_result(data: dict) -> tuple[str, str, str]:
    runs = int(data.get("runs", 0))
    clears = int(data.get("r10Clears", 0))
    rate = float(data.get("r10SuccessRate", 0.0))
    fate_uses = int(data.get("fateUses", 0))
    results = data.get("results", []) or []
    avg_life = sum(float(row.get("endLife", 0)) for row in results) / max(1, len(results))
    avg_gold = sum(float(row.get("endGold", 0)) for row in results) / max(1, len(results))
    avg_summons = sum(float(row.get("summons", 0)) for row in results) / max(1, len(results))
    failures = [row for row in results if not row.get("clearedR10", False)]
    failure_hp = [float(row.get("r10BossHealthRemaining01", -1)) for row in failures if float(row.get("r10BossHealthRemaining01", -1)) >= 0]
    avg_failure_hp = sum(failure_hp) / len(failure_hp) if failure_hp else 0.0

    if 0.60 <= rate <= 0.80:
        verdict = "통과"
        verdict_note = "5판 표본 해상도에서 목표 65~75%에 인접"
    elif rate > 0.80:
        verdict = "높음"
        verdict_note = "목표보다 쉬움"
    else:
        verdict = "낮음"
        verdict_note = "목표보다 어려움"

    result_text = f"R10 {clears}/{runs} ({rate:.0%}) / 운명 {fate_uses}/{runs}"
    detail = f"평균 종료 HP {avg_life:.1f}, 골드 {avg_gold:.1f}, 소환 {avg_summons:.1f}회"
    if failures:
        detail += f", 실패 {len(failures)}판 보스 잔여 HP 평균 {avg_failure_hp:.0%}"
    basis = f"{verdict_note}; 40배 안전속도·고정 시드·실제 R10 종료 판정"
    return result_text, verdict, detail + " / " + basis


def main() -> None:
    if not RESULT.exists():
        raise FileNotFoundError(RESULT)
    data = json.loads(RESULT.read_text(encoding="utf-8-sig"))
    if data.get("status") != "complete" or int(data.get("runs", 0)) != 5:
        raise RuntimeError("final 5-run result is not complete")

    base.main()
    wb = load_workbook(base.OUTPUT)
    latest = wb[base.LATEST_SHEET]
    validation = wb[base.VALIDATION_SHEET]
    marker_row = find_marker_row(latest, "9점 보완 조정 (2026-07-11)")
    for row in FINAL_LATEST_ROWS:
        upsert_latest(latest, marker_row, row)

    result_text, verdict, detail = summarize_result(data)
    base.upsert_validation(
        validation,
        "9점 보완 인간형 3전략 5판",
        [
            "9점 보완 인간형 3전략 5판",
            result_text,
            "BatchPlaytestResults/DefenseGame_Playtest5_Human3.json",
            verdict,
            detail,
            "summon-heavy / balanced / shop-save",
        ],
    )
    base.upsert_validation(
        validation,
        "R5~R9 일반 몬스터 완화",
        [
            "R5~R9 일반 몬스터 완화",
            "기본 수 -1",
            "RoundManager.cs",
            "적용",
            "초반 3소환 보호와 R10 전 압박 사이 중간값",
            "R5~R9",
        ],
    )
    base.upsert_validation(
        validation,
        "R10 첫 보스 압박",
        [
            "R10 첫 보스 압박",
            "HP x2.75 / 공격 x1.78 / 스킬 x1.45 / 지원 6기 / 최대 스턴 2명",
            "MonsterDatabase.cs / RoundManager.cs",
            "적용",
            "R10 한정; 일반 라운드·후속 보스 표는 유지",
            "Boss encounter 0",
        ],
    )

    errors = base.scan_formula_errors(wb)
    if errors:
        raise RuntimeError("formula errors: " + "; ".join(errors[:10]))
    wb.save(base.OUTPUT)
    shutil.copy2(base.OUTPUT, base.SOURCE)

    check = load_workbook(base.SOURCE, data_only=False)
    check_latest = check[base.LATEST_SHEET]
    check_validation = check[base.VALIDATION_SHEET]
    validation_row = next(
        row for row in range(1, check_validation.max_row + 1)
        if check_validation.cell(row, 1).value == "9점 보완 인간형 3전략 5판"
    )
    payload = {
        "source": str(base.SOURCE),
        "output": str(base.OUTPUT),
        "sheet_count": len(check.sheetnames),
        "latest_max_row": check_latest.max_row,
        "validation": [check_validation.cell(validation_row, col).value for col in range(1, 7)],
        "formula_errors": base.scan_formula_errors(check),
        "source_size": base.SOURCE.stat().st_size,
        "output_size": base.OUTPUT.stat().st_size,
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
